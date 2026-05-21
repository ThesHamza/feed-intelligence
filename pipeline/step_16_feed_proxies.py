"""PHASE 16 — Feed phosphate price proxies (public, free).

Feedinfo's actual MCP/MDCP/DCP assessments are paywalled. As a free, automated
proxy we use the World Bank "Pink Sheet" monthly commodity data, which includes
phosphate-chain prices that move directionally with feed phosphates:

  - Phosphate rock (FOB North Africa)  ← OCP's core upstream product
  - DAP diammonium phosphate (FOB US Gulf)
  - TSP triple superphosphate
  - Urea

IMPORTANT CAVEAT (shown in dashboard): these are fertilizer / upstream proxies,
NOT feed-grade phosphate prices. The correlation is directional, not exact —
Chinese producers now set the global feed phosphate floor independently of the
phosphoric acid / fertilizer complex. Use for trend direction, not absolute level.

The Pink Sheet is published monthly as an Excel workbook at a stable URL.

Usage: python -m pipeline.step_16_feed_proxies
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime, timezone

import requests

from . import config

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("feed_proxies")

# Stable World Bank Pink Sheet monthly Excel URL
PINK_SHEET_URL = "https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-Historical-Data-Monthly.xlsx"

# Series we extract (column header substrings as they appear in the 'Monthly Prices' sheet)
PROXY_SERIES = {
    "Phosphate rock": {"label": "Phosphate rock (FOB N. Africa)", "color": "#2563eb", "unit": "$/mt",
                       "note": "OCP core upstream product"},
    "DAP":            {"label": "DAP (FOB US Gulf)",             "color": "#16a34a", "unit": "$/mt",
                       "note": "Fertilizer proxy, directional"},
    "TSP":            {"label": "TSP",                           "color": "#ca8a04", "unit": "$/mt",
                       "note": "Phosphate fertilizer"},
    "Urea":           {"label": "Urea",                          "color": "#94a3b8", "unit": "$/mt",
                       "note": "Nitrogen reference"},
}

LOOKBACK_MONTHS = 36


def fetch_pink_sheet() -> dict | None:
    """Download and parse the World Bank Pink Sheet monthly workbook."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return None

    try:
        log.info("Downloading World Bank Pink Sheet…")
        resp = requests.get(PINK_SHEET_URL, timeout=60)
        resp.raise_for_status()
    except Exception as e:  # noqa: BLE001
        log.error("Download failed: %s", e)
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        log.error("Failed to parse workbook: %s", e)
        return None

    # The monthly data is in a sheet typically named 'Monthly Prices'
    sheet_name = None
    for name in wb.sheetnames:
        if "monthly" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row that contains commodity names (search first ~10 rows)
    header_row_idx = None
    for i, row in enumerate(rows[:12]):
        joined = " ".join(str(c) for c in row if c)
        if "Phosphate" in joined or "DAP" in joined:
            header_row_idx = i
            break
    if header_row_idx is None:
        log.error("Could not locate header row in Pink Sheet")
        return None

    headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]

    # Map each proxy series to its column index (first header containing the key)
    col_map = {}
    for key in PROXY_SERIES:
        for ci, h in enumerate(headers):
            if key.lower() in h.lower():
                col_map[key] = ci
                break

    if not col_map:
        log.error("No proxy columns matched. Headers sample: %s", headers[:15])
        return None

    log.info("Matched %d proxy columns: %s", len(col_map), list(col_map.keys()))

    # Data rows start after the header (+ unit row). Dates are usually in col 0 as 'YYYYMmm'
    series = {key: {"dates": [], "values": []} for key in col_map}
    data_rows = rows[header_row_idx + 1:]

    for row in data_rows:
        if not row or row[0] is None:
            continue
        date_label = str(row[0]).strip()
        # World Bank uses format like '2024M01'
        if "M" not in date_label:
            continue
        try:
            year, month = date_label.split("M")
            iso = f"{int(year):04d}-{int(month):02d}"
        except (ValueError, IndexError):
            continue

        for key, ci in col_map.items():
            if ci >= len(row):
                continue
            val = row[ci]
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            series[key]["dates"].append(iso)
            series[key]["values"].append(round(fval, 2))

    # Keep last LOOKBACK_MONTHS
    for key in series:
        series[key]["dates"] = series[key]["dates"][-LOOKBACK_MONTHS:]
        series[key]["values"] = series[key]["values"][-LOOKBACK_MONTHS:]

    return series


def build_output(series: dict) -> dict:
    proxies = []
    for key, meta in PROXY_SERIES.items():
        if key not in series or not series[key]["values"]:
            continue
        vals = series[key]["values"]
        dates = series[key]["dates"]
        latest = vals[-1]
        m3 = vals[-4] if len(vals) >= 4 else vals[0]
        m12 = vals[-13] if len(vals) >= 13 else vals[0]

        def pct(a, b):
            return round((a - b) / b * 100, 1) if b else None

        proxies.append({
            "key": key,
            "label": meta["label"],
            "color": meta["color"],
            "unit": meta["unit"],
            "note": meta["note"],
            "dates": dates,
            "values": vals,
            "latest": latest,
            "change_3m_pct": pct(latest, m3),
            "change_12m_pct": pct(latest, m12),
        })

    return {
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "source": "World Bank Commodity Markets (Pink Sheet), monthly",
        "caveat": "Fertilizer/upstream proxies, directional only — NOT feed-grade MCP/DCP/MDCP prices. Chinese producers now set the global feed phosphate floor independently.",
        "proxies": proxies,
    }


def main() -> None:
    series = fetch_pink_sheet()
    if not series:
        log.warning("No data fetched; existing feed_proxies.json (if any) left untouched.")
        return
    out = build_output(series)
    out_path = config.SITE_DATA_DIR / "feed_proxies.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info("Wrote %d proxy series → %s", len(out["proxies"]), out_path.name)


if __name__ == "__main__":
    main()
